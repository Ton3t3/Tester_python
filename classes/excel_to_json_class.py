from openpyxl import load_workbook

class ExccelToJson:
    def __init__(self):
        self.excel_url = ""

    def generate_json(self):
        workbook = load_workbook(self.excel_url)
        sheet = workbook.active
        rows = sheet.max_row

        ''' 0 -> Titulo test  1 -> Pregunta  2-5 -> Opciones  6 -> Respuesta Correcta  7 -> Foto bool  8 -> Foto Name '''
        questions = []
        options_array = ['a', 'b', 'c', 'd']
        for i in range(2, rows+1):
            question = {
                "test": str(sheet.cell(row = i, column = 1).value),
                "question": str(sheet.cell(row = i, column = 2).value),
                "options": [
                    {
                        "a": str(sheet.cell(row = i, column = 3).value),
                        "b": str(sheet.cell(row = i, column = 4).value),
                        "c": str(sheet.cell(row = i, column = 5).value),
                        "d": str(sheet.cell(row = i, column = 6).value)
                    }
                ],
                "solution": options_array[int(sheet.cell(row = i, column = 7).value)-1] if sheet.cell(row = i, column = 7).value in range(1,5) else str(sheet.cell(row = i, column = 7).value),
                "image": [
                    {
                        "img_bool": int(sheet.cell(row = i, column = 8).value),
                        "img_name": str(sheet.cell(row = i, column = 9).value) if bool(int(sheet.cell(row = i, column = 8).value)) else ""
                    }
                ]
            }
            questions.append(question)
        final_json = {"emp_details":questions}
        return final_json